import sqlite3
import datetime
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import webbrowser
import tempfile
import sys
import os
from main_window import MainWindow


class BorrowerExcel:
    def __init__(self, directory):
        connection = sqlite3.connect('munange.db')
        cursor = connection.cursor()
        query = ("SELECT customers.customer_id, customers.name, customers.gender, loans.loan_id, loans.amount, loans.loan_date, "
                 "loans.loan_deadline FROM customers JOIN loans ON customers.customer_id = loans.customer_no WHERE "
                 "loans.status='on going' ORDER BY name ASC")
        cursor.execute(query)
        borrowers = cursor.fetchall()
        # print(borrowers)
        global current
        self.borrowers_list_excel = []
        count = 0
        for borrower in borrowers:

            cursor.execute("SELECT amount, date FROM payments WHERE payment_id=?", (borrower[3], ))
            payments_info = cursor.fetchall()
            # print(payments_info)
            # working on the days remaining to deadline
            time_method = datetime(date.today().year, date.today().month, date.today().day)
            day = time_method.strftime("%d")
            month = time_method.strftime("%b")
            year = time_method.strftime("%Y")

            current = f'{day}-{month}-{year}'
            deadline = borrower[6]

            current_date = datetime.strptime(current, '%d-%b-%Y')
            deadline_date = datetime.strptime(deadline, '%d-%b-%Y')
            if current_date >= deadline_date:
                remaining_days = '0'
            else:
                remaining_days = str(deadline_date - current_date).split(',')[0]

            # working_with_getting_missed_days
            loan_date = borrower[5]

            days_list = []
            if current_date >= deadline_date:
                for day in range(30 - int(remaining_days.split()[0])):
                    loan_taken_date = datetime.strptime(loan_date, '%d-%b-%Y')
                    delta = timedelta(days=1)
                    date_ = loan_taken_date + delta
                    day = date_.strftime("%d")
                    month = date_.strftime("%b")
                    year = date_.strftime("%Y")
                    days_list.append(f'{day}-{month}-{year}')
                    loan_date = f'{day}-{month}-{year}'
            else:
                for day in range(30 - int(remaining_days.split()[0]) - 1):
                    loan_taken_date = datetime.strptime(loan_date, '%d-%b-%Y')
                    delta = timedelta(days=1)
                    date_ = loan_taken_date + delta
                    day = date_.strftime("%d")
                    month = date_.strftime("%b")
                    year = date_.strftime("%Y")
                    days_list.append(f'{day}-{month}-{year}')
                    loan_date = f'{day}-{month}-{year}'

            # print(days_list)
            self.paid_days = []
            current_amount = 0
            try:
                for payment in payments_info:
                    current_amount += int(payment[0])
                    self.paid_days.append(payment[1])
            except IndexError:
                pass

            missed_days = list(days for days in days_list if days not in self.paid_days)
            outstanding_balance = int(borrower[4]) - current_amount
            if current_date > deadline_date:
                remaining_days = 'Passed deadline'

            self.borrowers_list_excel.append((borrower[0], borrower[1], borrower[2],
                                  remaining_days, borrower[4], outstanding_balance))
            count += 1
        print(self.borrowers_list_excel)
        cursor.close()
        connection.close()

        wb = Workbook()
        # wb = load_workbook(filename='practice.xlsx')
        ws = wb.active

        title_font = Font(name='Calibri', size=14, bold=True)
        values_font = Font(name='Calibri', size=14)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15

        ws['A1'] = f'BORROWERS INFORMATION REPORT AS OF {current.upper()}'
        ws['A1'].font = title_font

        titles = ('S/No', 'ACCESS NO', 'NAME', 'GENDER', 'DAYS TO DEADLINE', 'AMOUNT', 'BALANCE')
        ws['A2'] = titles[0]
        ws['B2'] = titles[1]
        ws['C2'] = titles[2]
        ws['D2'] = titles[3]
        ws['E2'] = titles[4]
        ws['F2'] = titles[5]
        ws['G2'] = titles[6]

        title_values = ws['A2:G2']
        # print(title_values)
        for name in title_values[0]:
            name.font = title_font
        # ws[f'A2'] = 1
        ws.merge_cells('A1:G1')

        i = 3
        for borrower in self.borrowers_list_excel:
            ws[f'A{i}'] = i-2
            ws[f'B{i}'] = borrower[0]
            ws[f'C{i}'] = borrower[1]
            ws[f'D{i}'] = borrower[2]
            ws[f'E{i}'] = borrower[3]
            ws[f'F{i}'] = borrower[4]
            ws[f'G{i}'] = borrower[5]
            # ws.append([, student[1], student[2], student[3], student[4]], start_column=2)
            # print(ws)
            values = ws[f'A{i}:G{i}']
            for info in values[0]:
                info.font = values_font

            i += 1
        # Aligning column one
        for cell in ws['A']:
            cell.alignment = Alignment(horizontal='center')
        for cell in ws['B']:
            cell.alignment = Alignment(horizontal='center')

        if directory:
            wb.save(f"{directory}/Borrowers' Status.xlsx")

        else:

            # Save the workbook to a BytesIO object
            file_stream = BytesIO()
            wb.save(file_stream)
            # Write the BytesIO object to a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file.write(file_stream.getvalue())
                temp_file_path = temp_file.name

            # Open the temporary file with the default application (Excel)
            webbrowser.open(f"file://{temp_file_path}")

