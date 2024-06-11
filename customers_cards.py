import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import webbrowser
import tempfile
import sys
import os
from main_window import MainWindow


class CustomerExcel:
    def __init__(self, directory):
        connection = sqlite3.connect('munange.db')
        cursor = connection.cursor()
        cursor.execute("SELECT name, gender, phone, email, nin, district FROM customers ORDER BY name ASC")
        customers = cursor.fetchall()
        cursor.close()
        connection.close()

        wb = Workbook()
        # wb = load_workbook(filename='practice.xlsx')
        ws = wb.active

        title_font = Font(name='Calibri', size=14, bold=True)
        values_font = Font(name='Calibri', size=14)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 25
        title = ('S/No', 'NAME', 'GENDER', 'CONTACT', 'EMAIL', 'NIN', 'DISTRICT')
        ws.append(title)
        # print(ws['A1'])
        title_values = ws['A1:H1']
        # print(title_values)
        for name in title_values[0]:
            name.font = title_font
        # ws[f'A2'] = 1
        i = 2
        # print(students)
        for customer in customers:
            ws[f'A{i}'] = i-1
            ws[f'B{i}'] = customer[0]
            ws[f'C{i}'] = customer[1]
            ws[f'D{i}'] = customer[2]
            ws[f'E{i}'] = customer[3]
            ws[f'F{i}'] = customer[4]
            ws[f'G{i}'] = customer[5]
            # ws.append([, student[1], student[2], student[3], student[4]], start_column=2)
            # print(ws)
            values = ws[f'A{i}:G{i}']
            for info in values[0]:
                info.font = values_font

            i += 1
        # Aligning column one
        for cell in ws['A']:
            cell.alignment = Alignment(horizontal='center')

        if directory:
            wb.save(f'{directory}/All customers.xlsx')

        else:

            # Save the workbook to a BytesIO object
            file_stream = BytesIO()
            wb.save(file_stream)
            # Write the BytesIO object to a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file.write(file_stream.getvalue())
                temp_file_path = temp_file.name

            # Open the temporary file with the default application (Excel)
            webbrowser.open(f"file://{temp_file_path}")

